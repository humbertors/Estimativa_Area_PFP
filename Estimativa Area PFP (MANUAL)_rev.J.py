# Bibliotecas ----------------------------------------------------------------------------------------------------------
import os
import openpyxl
import re
import math
from math import pi
from tkinter.filedialog import askopenfilename  # NOVO

# Funções --------------------------------------------------------------------------------------------------------------
if True:

    def data_matrix(path, data_file_name, number_of_columns):
        file_path = os.path.join(path, data_file_name)
        data_file = open(file_path)
        data_string = data_file.read()
        data_file.close()
        data_regex = re.compile(r'[\w.()+-,]{1,100}')
        data_list = data_regex.findall(data_string)
        data_mat = []
        for k in range(int(len(data_list) / number_of_columns)):
            line = []
            for m in range(number_of_columns):
                line.append(0)
            data_mat.append(line)
        for k in range(len(data_mat)):
            for m in range(number_of_columns):
                data_mat[k][m] = data_list[k * number_of_columns + m]
        return data_mat


    def data_value(matrix, value, column, ref):
        for index in range(1, len(matrix)):
            if ref == 0 and matrix[index][0] == value:
                return float(matrix[index][column])
            elif ref == -1 and float(matrix[index][0]) > float(value):
                return float(matrix[index - 1][column])
            elif ref == 1 and float(matrix[index][0]) >= float(value):
                return float(matrix[index][column])

# Variáveis de entrada -------------------------------------------------------------------------------------------------
if True:
    path_name = askopenfilename(filetypes=(('Excel files', '*.xlsx'), ('All files', '*.*')))
    _dir, _name = os.path.split(path_name)
    t_jacket = 50  # espessura da manta [mm]
    t_jacket_mj = 25  # espessura da manta para fabricação dos mata-juntas [mm]
    b_jacket_mj = 100  # largura da manta para fabricação dos mata-juntas [mm]
    t_epoxi = 3  # espessura do revestimento de epóxi
    contingencia = 0.15  # fator de contingência

# Leitura dos arquivos e geração das matrizes de dados------------------------------------------------------------------
if True:
    file = openpyxl.load_workbook(path_name)

    sheet_name = file.sheetnames[0]
    sheet = file[sheet_name]

    number_lines = sheet.max_row
    result_column = sheet.max_column + 1

    data_dir = os.path.join(os.getcwd(), 'Tables')
    cap_matrix = data_matrix(data_dir, r'Cap_ASMEB16.9.txt', 4)
    tee_matrix = data_matrix(data_dir, r'T_ASMEB16.9.txt', 5)
    red_tee_matrix = data_matrix(data_dir, r'TRed_ASMEB16.9.txt', 3)
    do_flange_matrix = data_matrix(data_dir, r'diametro-flange_ASME_NORSOK_API.txt', 14)
    t_flange_matrix = data_matrix(data_dir, r'espessura-flange_ASME_NORSOK_API.txt', 14)
    len_flange_matrix = data_matrix(data_dir, r'comprimento-flangeWN_ASME_NORSOK_API.txt', 14)
    do_fig8_matrix = data_matrix(data_dir, r'diametro-fig8_ASME.txt', 8)
    t_fig8_matrix = data_matrix(data_dir, r'espessura-fig8_ASME.txt', 8)
    do_valve_matrix = data_matrix(data_dir, r'diametro-valvulas_CATALOGO.txt', 8)
    t_valve_matrix = data_matrix(data_dir, r'espessura-valvula_ASMEB16.34.txt', 8)
    len_valve_matrix = data_matrix(data_dir, r'comprimento-VES_API6D.txt', 7)

# Cálculos de área -----------------------------------------------------------------------------------------------------
if True:
    title = []
    for i in range(sheet.max_column):
        title.append(sheet.cell(row=1, column=i + 1).value)

    for i, tit in enumerate(title):
        if not tit:
            title[i] = i
            sheet.cell(row=1, column=i + 1).value = i

    type_column = title.index('TYPE') + 1
    nps1_column = title.index('FIRST_DIAM') + 1
    ri1_column = title.index('R_Int_Metro1') + 1  # Humberto: só para CAP
    # ro1_column = title.index('R_Ext_Metro1') + 1
    nps2_column = title.index('SECOND_DIAM') + 1
    # ri2_column = title.index('R_Int_Metro2') + 1
    # ro2_column = title.index('R_Ext_Metro2') + 1
    length_column = title.index('Comprimento [m]') + 1
    rating_column = title.index('RATING') + 1  # Humberto: pendente inserir valores
    description_column = title.index('DESCRIPTION') + 1
    spec_column = title.index('SPEC') + 1  # Humberto: pendente confirmar spec
    tag_piperun_column = title.index('TAG_PIPERUN') + 1
    material_category_column = title.index('MaterialCategory') + 1  # Humberto: não é usado?
    material_grade_column = title.index('MaterialGrade') + 1  # Humberto: só para PVC
    class_column = title.index('Class') + 1  # Humberto: usado só para válvulas?
    sub_class_column = title.index('SubClass') + 1  # Humberto: não é usado?
    tag_component_column = title.index('TAG_COMPONENT') + 1  # Humberto: usado para válvula controle e instrumentos

    elbow_list = ['E45', 'E455D', 'E45LR', 'E90', 'E905D', 'E90LR', 'E90SR', 'E904D']
    rating_list = ['150#', '300#', '600#', '900#', '1500#', '2500#', '10000psi']
    valve_list = ['Angle pressure relief valve', 'BALL', 'BALLP', 'CKNS', 'Double block and bleed valve, piping class',
                  'NEE', 'GATR', 'GLO', 'CKWF', 'Generic body valve', 'GATE']  # Humberto: GATE

    coef = [[4.0386, -0.170],
            [3.8189, -0.162],
            [3.6215, -0.158],
            [3.6651, -0.161],
            [2.6271, -0.104],
            [2.9894, -0.121],
            [3.8430, -0.214]]

    nps_diameter_inch_dict = {0.5: 0.84,
                              0.75: 1.05,
                              1: 1.315,
                              1.5: 1.9,
                              2: 2.375,
                              3: 3.5,
                              4: 4.5,
                              6: 6.625,
                              8: 8.625,
                              10: 10.75,
                              12: 12.75}

    specs_cs_list = ['B6', 'B9', 'B10', 'C10', 'E6', 'E10', 'E10N', 'F6X', 'F10', 'H10', 'H10N', 'H10PN', 'H30N',
                     'H30PN', 'B4A1']  # Humberto: B4A1
    specs_ss_list = ['B16', 'C16', 'E16N', 'H3N', 'E3X']

    tag_piperun_regex = re.compile(r'''
    (                       # GRUPO 0 - Tag completo da piperun
    ([0-9 .,;/"]{1,10})     # GRUPO 1 - Identificador do diâmetro nominal da linha
    -                       # Primeiro separador
    ([A-Za-z0-9]{1,10})     # GRUPO 2 - Identificador do fluido de processo
    -                       # Segundo separador
    ([A-Za-z0-9]{1,10})     # GRUPO 3 - Identificador do spec de tubulação
    -                       # Terceiro separador
    ([A-Za-z0-9]{1,10})     # GRUPO 4 - Número sequencial da linha
    ([- ]{1,6})?            # GRUPO 5 - Quarto separador (opcional)
    ([A-Za-z0-9]{1,10})?    # GRUPO 6 - Identificador do isolamento térmico da linha (opcional)
    ([A-Z0-9_-]{1,20})?     # GRUPO 7 - Identificador do módulo da linha (opcional)
    )
    ''', re.VERBOSE)

    sheet.cell(row=1, column=result_column).value = 'Tag Pipeline'
    sheet.cell(row=1, column=result_column + 1).value = 'ISOMETRICO'
    sheet.cell(row=1, column=result_column + 2).value = 'NPS PRINCIPAL'
    sheet.cell(row=1, column=result_column + 3).value = 'Comprimento equivalente [m]'
    sheet.cell(row=1, column=result_column + 4).value = 'Diametro equivalente [m]'
    sheet.cell(row=1, column=result_column + 5).value = 'Tipo de revestimento'
    sheet.cell(row=1, column=result_column + 6).value = 'Espessura do revestimento [mm]'
    sheet.cell(row=1, column=result_column + 7).value = 'Espessura do isolamento térmico [mm]'
    sheet.cell(row=1, column=result_column + 8).value = 'Area externa PFP atualizada [m2]'
    sheet.cell(row=1, column=result_column + 9).value = 'Quantidade de mata-juntas'
    sheet.cell(row=1, column=result_column + 10).value = 'Area PFP para mata-juntas em tubos [m2]'
    sheet.cell(row=1, column=result_column + 11).value = f'Area de PFP para contingência ' \
                                                         f'({contingencia * 100:.0f}%) [m2]'
    sheet.cell(row=1, column=result_column + 12).value = 'Area total de PFP [m2]'
    sheet.cell(row=1, column=result_column + 13).value = 'Percentual sobre o total geral [%]'
    sheet.cell(row=1, column=result_column + 14).value = 'Rating PFP'
    sheet.cell(row=1, column=result_column + 15).value = 'Seção a proteger'
    sheet.cell(row=1, column=result_column + 16).value = 'Obs EEAT'

    count = 0
    ctf_dict = {}
    for i in range(2, number_lines + 1):
        print(f'Current line: {i}')
        type_item = sheet.cell(row=i, column=type_column).value
        length = sheet.cell(row=i, column=length_column).value
        nps1 = sheet.cell(row=i, column=nps1_column).value
        nps2 = sheet.cell(row=i, column=nps2_column).value
        ri1 = sheet.cell(row=i, column=ri1_column).value
        # ro1 = sheet.cell(row=i, column=ro1_column).value
        if nps1:
            if nps1 > 12:
                do1 = nps1 * 0.0254
            else:
                d01 = nps_diameter_inch_dict.get(nps1, nps1) * 0.0254
                # do1 = nps_diameter_inch_dict[nps1] * 0.0254
        else:
            do1 = 0
        ro1 = do1 / 2

        if nps2:
            if nps2 > 12:
                do2 = nps2 * 0.0254
            else:
                d02 = nps_diameter_inch_dict.get(nps2, nps2) * 0.0254
                # do2 = nps_diameter_inch_dict[nps2] * 0.0254
        else:
            do2 = 0
        rating = sheet.cell(row=i, column=rating_column).value
        description = sheet.cell(row=i, column=description_column).value
        spec = sheet.cell(row=i, column=spec_column).value
        tag_piperun = sheet.cell(row=i, column=tag_piperun_column).value
        material_category = sheet.cell(row=i, column=material_category_column).value
        material_grade = sheet.cell(row=i, column=material_grade_column).value
        _class = sheet.cell(row=i, column=class_column).value
        sub_class = sheet.cell(row=i, column=sub_class_column).value
        tag_component = sheet.cell(row=i, column=tag_component_column).value or ''  # Humberto: or ''

        if tag_piperun:  # NOVO
            tag_piping = tag_piperun_regex.findall(tag_piperun)
            ctf = tag_piping[0][2] + '-' + tag_piping[0][4]
            if tag_piping[0][1] in ('1 1/2"', '1.5"'):  # Humberto in... (..., 1.5)
                diameter = 1.5
            elif tag_piping[0][1] in ('3/4"', '0.75"'):  # Humberto in... (..., 0.75)
                diameter = 0.75
            elif tag_piping[0][1] in ('1/2"', '0.5"'):  # Humberto in... (..., 0.5)
                diameter = 0.5
            else:
                diameter = int(tag_piping[0][1][:-1])
            if ctf in ctf_dict:
                ctf_dict[ctf] = max(diameter, ctf_dict[ctf])
            else:
                ctf_dict[ctf] = diameter

        # Comprimento e diâmetro equivalente de derivações: tês (T), tês de redução (TRB), tês oblíquos (LAT),
        # weldolets (WOL), sockolets (SOL), Couplings (CPL) e Half-couplings (CPLH)
        if type_item == 'T':
            if not length:
                C = data_value(tee_matrix, str(nps1), 1, 0)
                M = data_value(tee_matrix, str(nps2), 3, 0)
                length = (2 * C + M) / 1000
            length_eq = length - do1 / 2
            do_eq = do1

        elif type_item == 'TRB':
            if not length:
                C = data_value(red_tee_matrix, f'{nps1}x{nps1}x{nps2}', 1, 0)
                M = data_value(red_tee_matrix, f'{nps1}x{nps1}x{nps2}', 2, 0)
                length = (2 * C + M) / 1000
            length_eq = length - do1 / 2
            do_eq = do1

        elif type_item == 'LAT':
            length_eq = length - do1 / 2 / math.sin(30 * pi / 180)
            do_eq = do1

        elif type_item == 'WOL' or type_item == 'SOL' or type_item == 'CPL' or type_item == 'CPLH' or \
                type_item == 'Weldolet, heavy wall forged':
            length_eq = length
            do_eq = do2

        # Comprimento e diâmetro equivalente de Curvas
        elif type_item in elbow_list:
            angle = int(type_item[1:3])
            if type_item == 'E90' or type_item == 'E45':
                K = 1
            elif type_item[3:] == 'LR':
                K = 1.5
            elif type_item[3:] == 'SR':
                K = 1
            else:
                K = int(type_item[3])
            length_eq = angle * pi / 180 * (K * nps1 * 0.0254 + ro1)
            do_eq = do1

        # Comprimento e diâmetro equivalente de Caps
        elif type_item == 'CAP':
            if not length:
                E = data_value(cap_matrix, str(nps1), 1, 0) / 1000
                t_cap_max = data_value(cap_matrix, str(nps1), 2, 0) / 1000
                E1 = data_value(cap_matrix, str(nps1), 3, 0) / 1000
                t_cap = ro1 - ri1
                if t_cap <= t_cap_max:
                    length = E
                else:
                    length = E1
            length_eq = length + do1 / 4
            do_eq = do1

        # Comprimento e diâmetro equivalente de Flanges
        elif type_item == 'FWN' or type_item == 'FBLD' or type_item == 'FSW':

            if description:
                if '10000' in description or '5000' in description or 'API' in description:
                    standard = 'API'
                    flange_column_index = 13

                elif 'NORSOK' in description:
                    standard = 'NORSOK'
                    flange_column_index = rating_list.index(rating) + 7

                else:
                    standard = 'ASME'
                    flange_column_index = rating_list.index(rating) + 1
            else:
                standard = 'ASME'
                flange_column_index = rating_list.index(rating) + 1

            t_flange = data_value(t_flange_matrix, str(nps1), flange_column_index, 0)
            do_flange = data_value(do_flange_matrix, str(nps1), flange_column_index, 0) / 1000

            if not length:
                if type_item == 'FBLD':
                    length = t_flange / 1000
                elif type_item == 'FWN':
                    length = data_value(len_flange_matrix, str(nps1), flange_column_index, 0) / 1000
                else:
                    length = None

            length = max(length, (t_flange + 50) / 1000)
            length_eq = do_flange / do1 * (length + do_flange / 4)
            do_eq = do1

        # Comprimento e diâmetro equivalente de figuras 8
        elif type_item == 'BLSPO':
            fig8_column_index = rating_list.index(rating) + 1
            do_fig8 = data_value(do_fig8_matrix, str(nps1), fig8_column_index, 0) / 1000
            t_fig8 = data_value(t_fig8_matrix, str(nps1), fig8_column_index, 0)

            if not length:
                length = t_fig8 / 1000

            length_eq = 2 * do_fig8 / do1 * (length + do_fig8 / 4)
            do_eq = do1

        # Comprimento e diâmetro equivalente de válvulas
        elif type_item in valve_list or _class == 'Valves' or tag_component.startswith('FV-') \
                or tag_component.startswith('LV-') or tag_component.startswith('TV-') \
                or tag_component.startswith('PV-'):

            valve_column_index = rating_list.index(rating) + 1

            if nps1 < 2 or (rating == '900#' and nps1 > 36) or (rating == '1500#' and nps1 > 24) or \
                    (rating == '2500#' and nps1 > 12) or (rating == '10000psi' and nps1 > 12):
                di_cav_valve = 1.1 * do1 * 1000
                t_valve = data_value(t_valve_matrix, str(di_cav_valve), valve_column_index, 1)
                do_valve = di_cav_valve + 2 * t_valve
                coef_a = coef[rating_list.index(rating)][0]
                coef_b = coef[rating_list.index(rating)][1]
                factor = coef_a * do_valve ** coef_b
                do_valve = factor * do_valve / 1000

            else:
                do_valve = data_value(do_valve_matrix, str(nps1), valve_column_index, 0) * 0.0254

            do_eq = do_valve

            if not length:
                if nps1 == 0.75:
                    length = 0.28
                elif nps1 == 1:
                    length = 0.3
                elif nps1 == 1.5:
                    length = 0.39
                else:
                    length = data_value(len_valve_matrix, str(nps1), valve_column_index, 0) * 0.0254

            length_eq = length

        # Demais itens (Tubos retos, Reduções, Medidores de vazão)
        else:
            print(f'{i}. Comprimento equivalente do {type_item} calculado como demais itens.')
            length_eq = length
            do_eq = do1

        if not length_eq:
            print(f'{i}. O item "{description}" está com o comprimento equivalente zerado.')
            length_eq = 0

        # Escrevendo os resultados na planilha
        sheet.cell(row=i, column=result_column).value = ctf
        sheet.cell(row=i, column=result_column + 3).value = length_eq
        sheet.cell(row=i, column=result_column + 4).value = do_eq

    title = []
    for i in range(sheet.max_column):
        title.append(sheet.cell(row=1, column=i + 1).value)

    for i, tit in enumerate(title):
        if not tit:
            title[i] = i
            sheet.cell(row=1, column=i + 1).value = i

    ctf_column = title.index('Tag Pipeline') + 1
    do_eq_column = title.index('Diametro equivalente [m]') + 1
    length_eq_column = title.index('Comprimento equivalente [m]') + 1

    sum_a_pfp_total = 0

    for i in range(2, number_lines + 1):
        type_item = sheet.cell(row=i, column=type_column).value
        tag_piperun = sheet.cell(row=i, column=tag_piperun_column).value
        material_category = sheet.cell(row=i, column=material_category_column).value
        material_grade = sheet.cell(row=i, column=material_grade_column).value
        _class = sheet.cell(row=i, column=class_column).value
        sub_class = sheet.cell(row=i, column=sub_class_column).value
        ctf = sheet.cell(row=i, column=ctf_column).value
        do_eq = sheet.cell(row=i, column=do_eq_column).value
        length_eq = sheet.cell(row=i, column=length_eq_column).value
        nps1 = sheet.cell(row=i, column=nps1_column).value
        description = sheet.cell(row=i, column=description_column).value
        spec = sheet.cell(row=i, column=spec_column).value
        tag_component = sheet.cell(row=i, column=tag_component_column).value or ''  # Humberto: or ''

        if tag_piperun:  # NOVO
            nps_principal = ctf_dict[ctf]
        else:
            nps_principal = nps_diameter_inch_dict.get(nps1, None)
            if not nps_principal:
                nps_principal = max(float(nps1), float(nps2))

        # Definindo o tipo de revestimento (Jaqueta ou Epoxi)

        # Regra 4 – Eliminamos CAPs e pequenos elementos que estavam internos às tubulações, como placas de orifício;
        if (type_item == 'CAP' and nps1 and nps1 >= 2) or type_item == 'Orifice plate':
            type_coating = 'Sem revestimento'
            length_eq = 0
            obs_eeat = 'Premissa número 4 da UN-BS: "Eliminamos CAPs e pequenos elementos que estavam ' \
                       'internos às tubulações, como placas de orifício".'

        elif type_item == 'CKWF':
            type_coating = 'Sem revestimento'
            length_eq = 0
            obs_eeat = 'Válvulas wafer estão fora do escopo de PFP, conforme premissa informada pela EPS.'

        elif tag_component.startswith('SDV-') or tag_component.startswith('BDV-'):
            type_coating = 'Sem revestimento'
            length_eq = 0
            obs_eeat = 'Válvulas SDV e BDV estão fora do escopo de PFP, conforme premissa informada pela EPS.'

        elif tag_component.startswith('TIT-') or tag_component.startswith('PIT-') or tag_component.startswith('AIT-') \
                or tag_component.startswith('AX-') or tag_component.startswith('PDI-') \
                or tag_component.startswith('PDIT-') or tag_component.startswith('PI-') \
                or tag_component.startswith('TE-'):
            type_coating = 'Sem revestimento'
            length_eq = 0
            obs_eeat = 'Instrumentos estão fora do escopo de PFP, conforme premissa informada pela EPS.'

        # Regra 2 – Todos flanges e válvulas serão protegidos com Jaqueta;
        elif type_item in valve_list or _class == 'Valves' or tag_component.startswith('FV-') \
                or tag_component.startswith('LV-') or tag_component.startswith('TV-') \
                or tag_component.startswith('PV-') or type_item == 'FWN' or type_item == 'FBLD' or type_item == 'BLSPO':
            type_coating = 'Jaqueta'
            obs_eeat = 'Premissa número 2 da UN-BS: "Todo flange e válvula será protegido com Jaqueta".'

        # Regra 1 – Toda tubulação com isolamento térmico será protegida com Jaqueta;
        elif tag_piperun[-1].isalpha():
            type_coating = 'Jaqueta'
            obs_eeat = 'Premissa número 1 da UN-BS: "Toda tubulação com isolamento térmico será protegida' \
                       ' com Jaqueta".'

        # Regra 3 – Apareceram elementos de PVC como componentes de tubulação (com fluido HC) em alguns módulos.
        # Como o valor é pequeno, consideramos os mesmos como Jaqueta;
        elif material_grade and 'PVC' in material_grade:
            type_coating = 'Jaqueta'
            obs_eeat = 'Premissa número 3 da UN-BS: "Apareceram elementos de PVC como componentes de ' \
                       'tubulação (com fluido HC) em alguns módulos. Como o valor é pequeno, consideramos os mesmos ' \
                       'como Jaqueta".'

        # Regra extra (Guilherme Mariot) - Toda tubulação sem isolamento térmico em aço inox/duplex deverá ser protegida
        # com manta, pois tubulações não pintadas demandam uma preparação de superfície para aplicação do Epóxi
        elif spec in specs_ss_list:
            type_coating = 'Jaqueta'
            obs_eeat = 'Premissa extra definida pelo empreendimento (SRGE/SI-II/IES): tubulações em aço inox ou ' \
                       'duplex, independente do diâmetro, devem ser protegidas com manta, pois a aplicação de epóxi' \
                       'em tubulações não pintadas demandam uma grande preparação de superfície.'

        # Regra 5 - Toda tubulação sem isolamento térmico igual ou acima de 6” será protegida com Epóxi;
        elif not tag_piperun[-1].isalpha() and nps1 >= 6:
            type_coating = 'Epoxi'
            obs_eeat = 'Premissa número 5 da UN-BS: "Toda tubulação sem isolamento térmico igual ou acima ' \
                       'de 6” será protegida com Epóxi (exceto as de aço inox/duplex)".'

        # Regra 6 – As tubulações menores de 6” que forem em Aço Carbono deverão ser protegidas com Epóxi, devido risco
        # de corrosão sob isolamento (não tivemos esse caso na P-66);
        elif nps1 < 6 and spec in specs_cs_list:
            type_coating = 'Epoxi'
            obs_eeat = 'Premissa número 6 da UN-BS: "As tubulações menores de 6” que forem em Aço Carbono ' \
                       'deverão ser protegidas com Epóxi, devido risco de corrosão sob isolamento (não tivemos esse ' \
                       'caso na P-66)".'

        # Regra 7 – As demais tubulações menores que 6” deverão ser protegidas com Jaqueta.
        elif nps1 < 6:
            type_coating = 'Jaqueta'
            obs_eeat = 'Premissa número 7 da UN-BS: "As demais tubulações menores que 6” deverão ser ' \
                       'protegidas com Jaqueta".'

        # Demais casos
        else:
            print(f'{i}. Item {description} sem revestimento definido. Avaliar exclusão do escopo.')
            type_coating = 'Sem revestimento'
            obs_eeat = 'Item não definido. Avaliar exclusão do escopo.'

        sheet.cell(row=i, column=result_column + 2).value = nps_principal
        sheet.cell(row=i, column=result_column + 5).value = type_coating

        # Definindo a espessura do revestimento
        if type_coating == 'Jaqueta':
            t_coating = t_jacket
        elif type_coating == 'Epoxi':
            t_coating = t_epoxi
        else:
            t_coating = 0

        sheet.cell(row=i, column=result_column + 6).value = t_coating

        # Cálculo da área de revestimento com PFP
        if tag_piperun[-1].isalpha():
            t_insulation = 25
        else:
            t_insulation = 0
        a_pfp = pi * (do_eq + 2 * (t_coating + t_insulation) / 1000) * length_eq

        sheet.cell(row=i, column=result_column + 7).value = t_insulation
        sheet.cell(row=i, column=result_column + 8).value = a_pfp

        # Cálculo da área para fabricação de mata-juntas com manta de PFP em tubos
        if type_coating == 'Jaqueta' and type_item == 'PIPE':
            number_mj = length_eq // 1.2
            a_pfp_mj = number_mj * pi * (
                        do_eq + 2 * (t_jacket + t_jacket_mj + t_insulation) / 1000) * b_jacket_mj / 1000
        else:
            number_mj = 0
            a_pfp_mj = 0

        sheet.cell(row=i, column=result_column + 9).value = number_mj
        sheet.cell(row=i, column=result_column + 10).value = a_pfp_mj

        # Cálculo da área referente à contingência
        a_pfp_contingencia = a_pfp * contingencia
        sheet.cell(row=i, column=result_column + 11).value = a_pfp_contingencia

        # Cálculo da área total de PFP (área de revestimento + mata-juntas + contingência)
        a_pfp_total = a_pfp + a_pfp_mj + a_pfp_contingencia
        sheet.cell(row=i, column=result_column + 12).value = a_pfp_total
        sum_a_pfp_total += a_pfp_total

        sheet.cell(row=i, column=result_column + 16).value = obs_eeat

    title = []
    for i in range(sheet.max_column):
        title.append(sheet.cell(row=1, column=i + 1).value)

    for i, tit in enumerate(title):
        if not tit:
            title[i] = i
            sheet.cell(row=1, column=i + 1).value = i

    a_pfp_total_column = title.index('Area total de PFP [m2]') + 1

    for i in range(2, number_lines + 1):
        a_pfp_total = sheet.cell(row=i, column=a_pfp_total_column).value
        sheet.cell(row=i, column=result_column + 13).value = a_pfp_total / sum_a_pfp_total * 100

# Salvando os resultados -----------------------------------------------------------------------------------------------
# os.chdir(path_name)
new_name = 'Resultado_' + _name
new_path = os.path.join(_dir, new_name)
file.save(new_path)
